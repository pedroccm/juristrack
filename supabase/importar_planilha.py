#!/usr/bin/env python3
"""
JurisTrack - Importador de Planilha
=====================================
Lê a planilha de acompanhamento, cria usuários no Supabase Auth
e importa todos os processos com seus responsáveis.

Uso:
    python importar_planilha.py --planilha "../../judi/docs/Acompanhamento Diário 12.03.26.xlsx"
    python importar_planilha.py --planilha "..." --dry-run
"""

import os, sys, re, unicodedata, argparse, json, time
if sys.platform == "win32":
    os.environ.setdefault("PYTHONUTF8", "1")
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests
import openpyxl
from datetime import datetime

# ─── Configuração Supabase ────────────────────────────────────────────────────
SUPABASE_URL      = "https://gzhzdsahxgzlxxttsakc.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imd6aHpkc2FoeGd6bHh4dHRzYWtjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMyNTQwMjMsImV4cCI6MjA4ODgzMDAyM30.RPzPNmVS6X3BMRXm-HRNZPlAa8HmK2Upc0bzrkmGDls"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imd6aHpkc2FoeGd6bHh4dHRzYWtjIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MzI1NDAyMywiZXhwIjoyMDg4ODMwMDIzfQ.hR5feJmGxq1NE1SPfhvDEOwUJwDXXDOQCCkMkjHm-JE"

SENHA_TEMPORARIA = "Juris@2026"
EMAIL_DOMINIO    = "juristrack.com"

# ─── Índices da planilha ──────────────────────────────────────────────────────
COL_RESPONSAVEL   = 0
COL_PASTA         = 1
COL_CLIENTE       = 2
COL_DISCUSSAO     = 3
COL_NUMERO        = 4
COL_DATA          = 7
COL_ANDAMENTO     = 8
COL_DATA_ANT      = 10
COL_PLATAFORMA    = 15

RE_CNJ = re.compile(r'\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}')


# ─── Utilitários ─────────────────────────────────────────────────────────────

def slugify(nome: str) -> str:
    """
    'Antônio' → 'antonio'
    'Britto / Pedro' → não usado diretamente (split antes)
    """
    nome = unicodedata.normalize("NFD", nome)
    nome = "".join(c for c in nome if unicodedata.category(c) != "Mn")  # remove acentos
    nome = nome.lower().strip()
    nome = re.sub(r"[^a-z0-9]", "", nome)  # só letras e números
    return nome


def split_responsaveis(raw: str) -> list[str]:
    """
    'Britto / Pedro' → ['Britto', 'Pedro']
    'Antônio / Eduardo' → ['Antônio', 'Eduardo']
    'Ludmilla' → ['Ludmilla']
    'Controladoria SP / Giulia' → ['Controladoria SP', 'Giulia']
    """
    if not raw or raw.strip() in ("-", "N/A", ""):
        return []
    partes = [p.strip() for p in raw.split("/")]
    return [p for p in partes if p and p not in ("-", "N/A")]


def normalizar_numero(raw: str) -> list[str]:
    if not raw:
        return []
    matches = RE_CNJ.findall(str(raw))
    return [m.replace("-", "").replace(".", "") for m in matches]


def celula_str(v) -> str:
    return str(v).strip() if v is not None else ""


def celula_data(v):
    if isinstance(v, datetime):
        return v.isoformat()
    return None


# ─── Supabase helpers ─────────────────────────────────────────────────────────

class SupabaseClient:
    def __init__(self, service_key: str):
        self.base     = SUPABASE_URL
        self.anon_key = SUPABASE_ANON_KEY
        self.svc_key  = service_key
        self.headers_anon = {"apikey": SUPABASE_ANON_KEY, "Authorization": f"Bearer {SUPABASE_ANON_KEY}"}
        self.headers_svc  = {"apikey": service_key, "Authorization": f"Bearer {service_key}", "Content-Type": "application/json"}

    def criar_usuario_auth(self, email: str, nome: str, role: str) -> str | None:
        """Cria usuário via Admin API e insere em jt_usuarios. Retorna o UUID."""
        r = requests.post(
            f"{self.base}/auth/v1/admin/users",
            headers=self.headers_svc,
            json={
                "email": email,
                "password": SENHA_TEMPORARIA,
                "email_confirm": True,
                "user_metadata": {"nome": nome, "role": role},
            }
        )
        if r.status_code not in (200, 201):
            print(f"  ⚠️  Erro ao criar auth user {email}: {r.status_code} {r.text[:200]}")
            return None

        uid = r.json().get("id")
        if not uid:
            return None

        # Insere diretamente em jt_usuarios (não depende do trigger)
        ins = requests.post(
            f"{self.base}/rest/v1/jt_usuarios",
            headers={**self.headers_svc, "Prefer": "return=minimal"},
            json={"id": uid, "nome": nome, "email": email, "role": role, "ativo": True},
        )
        if ins.status_code not in (200, 201):
            print(f"  ⚠️  Auth criado mas falhou jt_usuarios {email}: {ins.status_code} {ins.text[:100]}")

        return uid

    def buscar_usuario_por_email(self, email: str) -> str | None:
        """Retorna UUID se o usuário já existe."""
        r = requests.get(
            f"{self.base}/rest/v1/jt_usuarios?email=eq.{email}&select=id",
            headers={**self.headers_svc, "Content-Type": "application/json"},
        )
        data = r.json()
        if data:
            return data[0]["id"]
        return None

    def inserir(self, tabela: str, dados: dict | list, ignore_conflict: bool = False) -> dict | None:
        prefer = "return=representation"
        if ignore_conflict:
            prefer = "return=representation,resolution=ignore-duplicates"
        for tentativa in range(3):
            try:
                r = requests.post(
                    f"{self.base}/rest/v1/{tabela}",
                    headers={**self.headers_svc, "Prefer": prefer},
                    json=dados,
                    timeout=30,
                )
                if r.status_code in (200, 201):
                    result = r.json()
                    return result[0] if isinstance(result, list) and result else result
                if r.status_code == 409:
                    return None  # conflito esperado, ignora
                print(f"  ⚠️  Erro ao inserir em {tabela}: {r.status_code} {r.text[:200]}")
                return None
            except Exception as e:
                if tentativa < 2:
                    time.sleep(2)
                else:
                    print(f"  ⚠️  Falha após 3 tentativas em {tabela}: {e}")
        return None


# ─── Leitura da planilha ──────────────────────────────────────────────────────

def ler_planilha(path: str) -> list[dict]:
    wb    = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows  = []
    sheets = ["Acompanhamento Diário", "2X POR SEMANA "]

    for sheet_name in sheets:
        ws = None
        for sn in wb.sheetnames:
            if sn.strip().lower().startswith(sheet_name.strip().lower()[:10]):
                ws = wb[sn]; break
        if not ws:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            numero_raw  = celula_str(row[COL_NUMERO] if len(row) > COL_NUMERO else "")
            plataforma  = celula_str(row[COL_PLATAFORMA] if len(row) > COL_PLATAFORMA else "").upper()
            responsavel = celula_str(row[COL_RESPONSAVEL] if len(row) > COL_RESPONSAVEL else "")

            if not numero_raw:
                continue

            numeros = normalizar_numero(numero_raw)
            if not numeros:
                continue

            rows.append({
                "responsaveis_raw": responsavel,
                "pasta":            celula_str(row[COL_PASTA] if len(row) > COL_PASTA else ""),
                "cliente":          celula_str(row[COL_CLIENTE] if len(row) > COL_CLIENTE else ""),
                "discussao":        celula_str(row[COL_DISCUSSAO] if len(row) > COL_DISCUSSAO else ""),
                "numero_cnj":       numero_raw.strip(),  # mantém formatado na tabela
                "numeros_limpos":   numeros,
                "plataforma":       plataforma,
                "andamento_atual":  celula_str(row[COL_ANDAMENTO] if len(row) > COL_ANDAMENTO else ""),
                "data_andamento":   celula_data(row[COL_DATA] if len(row) > COL_DATA else None),
            })

    print(f"Planilha lida: {len(rows)} processos encontrados.")
    return rows


# ─── Extrai e cria usuários únicos ────────────────────────────────────────────

def criar_usuarios(rows: list[dict], sb: SupabaseClient, dry_run: bool) -> dict[str, str]:
    """
    Retorna dict: slug → uuid
    Ex: {"antonio": "uuid-...", "ludmilla": "uuid-..."}
    """
    # Coleta todos os nomes únicos
    nomes_unicos: dict[str, str] = {}  # slug → nome original mais completo

    for row in rows:
        for nome in split_responsaveis(row["responsaveis_raw"]):
            slug = slugify(nome)
            if not slug:
                continue
            # Prefere o nome com mais caracteres (mais completo)
            if slug not in nomes_unicos or len(nome) > len(nomes_unicos[slug]):
                nomes_unicos[slug] = nome

    print(f"\nUsuários únicos encontrados: {len(nomes_unicos)}")
    for slug, nome in sorted(nomes_unicos.items()):
        email = f"{slug}@{EMAIL_DOMINIO}"
        print(f"  {nome:30s} → {email}")

    if dry_run:
        print("\n[DRY-RUN] Nenhum usuário criado.")
        return {slug: f"dry-run-uuid-{slug}" for slug in nomes_unicos}

    slug_to_uuid: dict[str, str] = {}

    for slug, nome in nomes_unicos.items():
        email = f"{slug}@{EMAIL_DOMINIO}"

        # Verifica se já existe
        existing = sb.buscar_usuario_por_email(email)
        if existing:
            print(f"  ✓ {email} já existe ({existing[:8]}...)")
            slug_to_uuid[slug] = existing
            continue

        uuid = sb.criar_usuario_auth(email, nome, "advogado")
        if uuid:
            print(f"  ✅ Criado: {email} ({uuid[:8]}...)")
            slug_to_uuid[slug] = uuid
        else:
            print(f"  ❌ Falhou: {email}")

        time.sleep(0.2)  # rate limit

    return slug_to_uuid


# ─── Importa processos ────────────────────────────────────────────────────────

def importar_processos(rows: list[dict], slug_to_uuid: dict[str, str], sb: SupabaseClient, dry_run: bool):
    total    = 0
    pulados  = 0
    erros    = 0

    for row in rows:
        # Monta número CNJ formatado (usa o primeiro se houver múltiplos)
        numero_display = row["numero_cnj"]
        # Pega número limpo → reformata para exibição padrão
        if row["numeros_limpos"]:
            n = row["numeros_limpos"][0]
            # Reformata: NNNNNNN-DD.AAAA.J.TT.OOOO
            if len(n) == 20:
                numero_display = f"{n[:7]}-{n[7:9]}.{n[9:13]}.{n[13]}.{n[14:16]}.{n[16:]}"

        responsaveis_slugs = [slugify(n) for n in split_responsaveis(row["responsaveis_raw"]) if slugify(n)]
        responsaveis_ids   = [slug_to_uuid[s] for s in responsaveis_slugs if s in slug_to_uuid]

        if not responsaveis_ids and responsaveis_slugs:
            print(f"  ⚠️  Responsáveis não mapeados: {row['responsaveis_raw']} → pulando")
            pulados += 1
            continue

        if dry_run:
            print(f"  [DRY-RUN] {numero_display} | {row['cliente'][:30]} | {row['plataforma']} | resp: {responsaveis_slugs}")
            total += 1
            continue

        # Insere o processo
        processo = sb.inserir("jt_processos", {
            "numero_cnj":     numero_display,
            "tribunal":       row["plataforma"],
            "cliente":        row["cliente"],
            "pasta":          row["pasta"] or None,
            "discussao":      row["discussao"] or None,
            "andamento_atual": row["andamento_atual"] or None,
            "data_andamento": row["data_andamento"],
            "status":         "ativo",
        })

        if not processo:
            erros += 1
            continue

        # Insere relacionamentos processo ↔ responsáveis
        for uid in responsaveis_ids:
            sb.inserir("jt_processo_responsaveis", {
                "processo_id": processo["id"],
                "usuario_id":  uid,
            }, ignore_conflict=True)

        time.sleep(0.1)

        # Insere andamento inicial no histórico
        if row["andamento_atual"]:
            sb.inserir("jt_andamentos", {
                "processo_id":    processo["id"],
                "descricao":      row["andamento_atual"],
                "data_andamento": row["data_andamento"] or datetime.now().isoformat(),
                "detectado_em":   datetime.now().isoformat(),
            })

        total += 1
        if total % 20 == 0:
            print(f"  ... {total} processos importados")

    print(f"\n✅ Importação concluída:")
    print(f"   Processos importados : {total}")
    print(f"   Pulados              : {pulados}")
    print(f"   Erros                : {erros}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--planilha", required=True)
    parser.add_argument("--dry-run",  action="store_true")
    args = parser.parse_args()

    sb = SupabaseClient(SUPABASE_SERVICE_KEY)  # type: ignore

    print("=" * 60)
    print("  JurisTrack — Importador de Planilha")
    print("=" * 60)

    rows = ler_planilha(args.planilha)
    slug_to_uuid = criar_usuarios(rows, sb, args.dry_run)

    print(f"\nImportando {len(rows)} processos...")
    importar_processos(rows, slug_to_uuid, sb, args.dry_run)


if __name__ == "__main__":
    main()
